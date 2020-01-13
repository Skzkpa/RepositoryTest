import logging
import warnings

import openpyxl
from commons import dump, process, save_yaml_file


warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=UserWarning)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def process_worksheet(ws):
    header = []
    data_set = []

    for row in ws.iter_rows(max_row=1):
        for cell in row:
            if cell.value:
                header.append(cell.value.lower().replace(' ', '_'))
    for row in ws.iter_rows(min_row=2):
        data = {}
        for idx, cell in enumerate(row):
            if cell.value:
                data[header[idx]] = cell.value
        data_set.append(data)
    return data_set


def process_excel_file(file):
    wb = openpyxl.load_workbook(file)
    cfg = {x['area']: x['velocity'] for x in process_worksheet(wb.get_sheet_by_name('teams'))}
    logging.info("Config obtained")
    for ws in wb.worksheets:
        area = ws.title.lower()
        if area.startswith('arkusz') or area.startswith('sheet') or area.startswith('common'):
            continue

        data_set = process_worksheet(ws)

        if area == "teams":
            output = dump(data_set)
        else:
            output = process(data_set, cfg.get(area, 1))
        save_yaml_file(output, area)
        logging.info(f"{area} Done")



process_excel_file('excel.xlsx')

logging.info(f"Finished")
